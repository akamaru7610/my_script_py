import os
import sys
from pathlib import Path
import openpyxl
import json

APP_HOME = os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".." ))
sys.path.append(APP_HOME)
PROG_NAME = os.path.splitext(os.path.basename(__file__))[0]

from lib.my_logger import MyLogger
from conf.my_batch_conf import MyBatchConf

MY_LOGGER = MyLogger.get_my_logger(APP_HOME, PROG_NAME)

def cmd():
    """
    """
    try:
        MY_LOGGER.info("start")

        # data_only=True は数値のみを取得する
        wb = openpyxl.load_workbook(MyBatchConf.target_excel, data_only=True)

        for sheet in wb:
            if sheet.title in MyBatchConf.no_chk_sheet:
                continue
            MY_LOGGER.info("sheet:{}".format(sheet.title))

            number_arr = get_number(sheet)
            name_arr   = get_name(sheet)
            digit_arr  = get_digit(sheet)
            data = get_out_content(number_arr, name_arr, digit_arr)
            # MY_LOGGER.info(json.dumps(data, ensure_ascii=False, indent=2))
            out_file_path = os.path.join(MyBatchConf.out_json_dir, (sheet.title + "_2.json"))
            out_setting_json(out_file_path, data)

    except Exception as e:
        MY_LOGGER.exception(e)
        sys.exit(1)
    else:
        MY_LOGGER.info("end")

def get_number(sheet):
    """
     №を取得します。
    """
    number_arr = []
    for row in sheet.iter_rows(min_row = MyBatchConf.content_row, min_col = MyBatchConf.number_col, max_col=MyBatchConf.number_col):
        for cell in row:
            if cell.value is None:
                continue
            number_arr.append(cell.value)
    return number_arr

def get_name(sheet):
    """
     項目名を取得します。
    """
    name_arr = []
    for row in sheet.iter_rows(min_row = MyBatchConf.content_row, min_col = MyBatchConf.name_col, max_col=MyBatchConf.name_col):
        for cell in row:
            if cell.value is None:
                continue
            name_arr.append(cell.value)
    return name_arr

def get_digit(sheet):
    """
     桁数を取得します。
    """
    digit_arr = []
    for row in sheet.iter_rows(min_row = MyBatchConf.content_row, min_col = MyBatchConf.digit_col, max_col = MyBatchConf.digit_col):
        for cell in row:
            if cell.value is None:
                continue
            digit_arr.append(cell.value)
    return digit_arr

def get_out_content(number_arr, name_arr, digit_arr):
    """
     出力内容を取得します。
    """
    data = dict()
    data["name"] = name_arr
    for num in number_arr:
        data["num"+str(num)] = digit_arr[num-1]
    return data

def out_setting_json(file_path, data):
    """
     設定情報をjson形式で出力します。
    """
    with open(file_path, mode='wt', encoding='utf-8') as file:
        json.dump(data, file, ensure_ascii=False, indent=2)

if __name__ == '__main__':
    cmd()
