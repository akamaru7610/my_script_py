import os
import sys
from pathlib import Path
import openpyxl
import json

APP_HOME = os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".." ))
sys.path.append(APP_HOME)
PROG_NAME = os.path.splitext(os.path.basename(__file__))[0]

from lib.my_logger import MyLogger
from conf.my_batch_conf import MyBatchConf

MY_LOGGER = MyLogger.get_my_logger(APP_HOME, PROG_NAME)

def cmd():
    """
     60_外部インターフェイスレイアウト定義書をもとに、設定情報をJSON形式で出力します。
     ■設定情報
       1. シート名
       2. 項目数
       3. 桁数の合計
    """
    try:
        MY_LOGGER.info("start")

        # data_only=True は数値のみを取得する
        wb = openpyxl.load_workbook(MyBatchConf.target_excel, data_only=True)

        for sheet in wb:
            if sheet.title in MyBatchConf.no_chk_sheet:
                continue
            MY_LOGGER.info("sheet:{}".format(sheet.title))

            data = dict()
            data['name']      = sheet.title
            data['item_cnt']  = get_item_cnt(sheet)
            data['sum_digit'] = get_sum_digit(sheet)
            # MY_LOGGER.info(json.dumps(data, ensure_ascii=False, indent=2))
            out_file_path = os.path.join(MyBatchConf.out_json_dir, (sheet.title + "_1.json"))
            out_setting_json(out_file_path, data)

    except Exception as e:
        MY_LOGGER.exception(e)
        sys.exit(1)
    else:
        MY_LOGGER.info("end")

def get_item_cnt(sheet):
    """
     項目数を取得します。
    """
    item_cnt = 0
    for row in sheet.iter_rows(min_row = MyBatchConf.content_row, min_col = MyBatchConf.number_col, max_col=MyBatchConf.number_col):
        for cell in row:
            if cell.value is None:
                continue
            item_cnt += 1
    return item_cnt

def get_sum_digit(sheet):
    """
     桁数の合計を取得します。
    """
    sum_digit = 0
    for row in sheet.iter_rows(min_row = MyBatchConf.content_row, min_col = MyBatchConf.digit_col, max_col=MyBatchConf.digit_col):
        for cell in row:
            if cell.value is None:
                continue
            sum_digit += cell.value
    # 改行文字をプラスする
    sum_digit += 1
    return sum_digit

def out_setting_json(file_path, data):
    """
     設定情報をjson形式で出力します。
    """
    with open(file_path, mode='wt', encoding='utf-8') as file:
        json.dump(data, file, ensure_ascii=False, indent=2)

if __name__ == '__main__':
    cmd()
